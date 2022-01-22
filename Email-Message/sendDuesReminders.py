#!/usr/bin/python3

import sys
import openpyxl
import smtplib

# open spreadsheet and get latest dues status
wb = openpyxl.loa_workbook('spreadsheet.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# check each member's payment status.
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment =! 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email    # dictionary of all members which we need to send notification mail

# log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.elho()
smtpObj.starttls()
smtpObj.login('yaser.amini67@gmail.com', sys.argv[1])   # sys.argv[1] is the password of gmail account


# send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'" % (latestMonth, name, latestMonth)

    print("Sending email to %s ..." % email)
    sendmailStatus = smtpObj.sendmail('yaser.amini67@gmail.com', email, body)

    if sendmailStatus != {}:
        print("There was a problem sending email to %s: %s" % ( email, sendmailStatus))

smtpObj.quit()        
